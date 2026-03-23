# -*- coding: utf-8 -*-
"""
将"企业服务数据新表.xlsx" Sheet 1.1 转换为 TriCenter 调研导入模板格式
生成的文件可直接通过 /api/survey-excel/import 接口导入

用法: python scripts/convert_enterprise_data.py
输入: docs/data/企业服务数据新表.xlsx (Sheet 1.1)
输出: docs/data/转换后_调研导入数据.xlsx
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import re
import os

# ==================== 配置 ====================

INPUT_FILE = os.path.join('docs', 'data', '企业服务数据新表.xlsx')
OUTPUT_FILE = os.path.join('docs', 'data', '转换后_调研导入数据.xlsx')
SOURCE_SHEET = '1.1 常州市产业带调研'

# 区域映射：新表简称 → 系统全称
DISTRICT_MAP = {
    '新北': '新北区', '武进': '武进区', '天宁': '天宁区',
    '钟楼': '钟楼区', '经开': '经开区', '金坛': '金坛区', '溧阳': '溧阳市',
    '新北区': '新北区', '武进区': '武进区', '天宁区': '天宁区',
    '钟楼区': '钟楼区', '经开区': '经开区', '金坛区': '金坛区', '溧阳市': '溧阳市',
}

# 企业类型映射
TYPE_MAP = {
    '工厂': '生产型', '贸易': '贸易型', '贸易公司': '贸易型',
    '工贸一体': '工贸一体', '生产型': '生产型', '贸易型': '贸易型',
}

# 行业别名 → 数据库标准名称映射
INDUSTRY_ALIAS_MAP = {
    '汽配': '汽车零部件', '汽车配件': '汽车零部件', '汽车部件': '汽车零部件',
    '汽车配套工具': '汽车零部件', '汽车制造': '汽车零部件',
    '纺织': '纺织服装', '纺织品': '纺织服装', '纺织业': '纺织服装',
    '纺织服装类': '纺织服装', '纺织科技': '纺织服装', '纺机': '纺织服装',
    '针纺织品': '纺织服装', '服装': '纺织服装', '服装制造': '纺织服装',
    '服装贸易进出口': '纺织服装', '纺织服装': '纺织服装',
    '医疗': '医疗器械', '医疗行业': '医疗器械', '制药装备': '医疗器械',
    '机械': '机械设备', '机械设备制造': '机械设备', '干燥设备': '机械设备',
    '干燥业': '机械设备', '通用设备': '机械设备', '35专用设备制造业': '机械设备',
    '通信设备': '机械设备',
    '电子': '电子产品', '电子电工': '电子产品', '电子机械': '电子产品',
    '电机行业': '电子产品', '3c': '电子产品', '3c电子': '电子产品',
    '五金': '五金建材', '五金工具': '五金建材', '五金工具制造': '五金建材',
    '建材': '五金建材', '建筑': '五金建材', '建筑工程': '五金建材',
    '装饰材料': '五金建材', '钣金制造': '五金建材', '高压配电': '五金建材',
    '齿轮': '五金建材',
    '灯具': '照明灯具', '灯具照明': '照明灯具', '照明': '照明灯具',
    '家居': '家居用品', '家具': '家居用品', '家居园艺': '家居用品',
    '智能家居': '家居用品', '厨具': '家居用品', '日用百货': '家居用品',
    '办公用品': '家居用品',
    '化工': '化工材料', '塑料包装': '化工材料', '新材料': '化工材料',
    '光伏': '新能源', '光伏行业': '新能源',
    '扳机': '电动工具', '箱包': '箱包皮具',
    '贸易': '综合贸易', '工贸一体': '综合贸易', '卖家': '综合贸易',
    '批发外贸': '综合贸易', '批发商': '综合贸易',
    '批发': '批发零售', '批发业': '批发零售', '批发和零售业': '批发零售',
    '批发零售业': '批发零售', '多品类': '批发零售',
    '电商': '电商运营', '跨境卖家': '电商运营', '跨境电商': '电商运营',
    '互联网销售': '电商运营',
    '进出口': '进出口代理', '艺术品进出口': '进出口代理', '技术进出口': '进出口代理',
    '软件': '信息技术', '信息工程': '信息技术', '信息传输软件业': '信息技术',
    '技术服务': '信息技术', '人工智能': '信息技术', '智能机器人': '信息技术',
    '高新技术业': '信息技术', '互联网信息': '信息技术',
    '海路运输': '物流仓储', '金融': '金融服务',
    '广告': '营销推广', '展示器材': '营销推广',
    '制造': '其他制造', '制造业': '其他制造', '工业制造': '其他制造',
    '工业品': '其他制造', '生产制造': '其他制造', '生产加工': '其他制造',
    '印刷': '其他制造', '文化': '其他制造',
    '专业设计服务': '咨询服务',
    # 复合名称和剩余映射
    '刀具': '五金建材', '轻工业': '其他制造', '机电行业': '机械设备',
    '成人用品': '其他制造',
    '汽车零部件，机械零部件，轴承': '汽车零部件',
    '新能源航空动力系统 / 电机制造': '新能源',
    '环保设备 / 净化科技': '机械设备',
    '计算机信息技术服务 / 系统集成': '信息技术',
    '建筑材料 / 化工产品 / 合成材料': '五金建材',
    '互联网 / 电子商务': '电商运营',
    '先进制造 / 电机及控制器制造': '机械设备',
    '服装贸易': '纺织服装',
    '建筑材料': '五金建材',
    '服务业': '其他制造',
}


def normalize_industry(val):
    """将原始行业名称标准化为数据库分类名称"""
    if not val:
        return ''
    val = str(val).strip()
    if val in ('.', ''):
        return ''
    mapped = INDUSTRY_ALIAS_MAP.get(val)
    if mapped:
        return mapped
    # 处理含斜杠的复合名称，取第一个关键词
    first = re.split(r'[/、，\s]', val)[0].strip()
    if first != val and first in INDUSTRY_ALIAS_MAP:
        return INDUSTRY_ALIAS_MAP[first]
    return val


# 是/否标准化
def normalize_yes_no(val):
    if not val:
        return ''
    val = str(val).strip().upper()
    if val in ('是', 'Y', 'YES', '有', '1', 'TRUE'):
        return '是'
    if val in ('否', 'N', 'NO', '无', '0', 'FALSE'):
        return '否'
    return ''


def clean_name(name):
    """清理企业名称：只去除首尾空格，保留括号内备注信息"""
    if not name:
        return ''
    return str(name).strip()


def safe_str(val):
    """安全转字符串，None → ''"""
    if val is None:
        return ''
    return str(val).strip()


def parse_revenue_range(val):
    """将营业额数字转换为系统选项区间（万元）"""
    if not val:
        return ''
    val = str(val).strip()
    # 尝试提取数字
    num = None
    if '亿' in val:
        m = re.search(r'([\d.]+)', val)
        if m:
            num = float(m.group(1)) * 10000  # 转万元
    elif '万' in val:
        m = re.search(r'([\d.]+)', val)
        if m:
            num = float(m.group(1))
    else:
        m = re.search(r'([\d.]+)', val)
        if m:
            num = float(m.group(1))
            # 如果数字很大，可能已经是万元
            if num < 100:
                num = num * 10000  # 假设是亿

    if num is None:
        return ''

    # 映射到系统选项
    if num < 200:
        return '200万以下'
    elif num < 500:
        return '200-500万'
    elif num < 1000:
        return '500-1000万'
    elif num < 5000:
        return '1000-5000万'
    elif num < 10000:
        return '5000万-1亿'
    else:
        return '1亿以上'


def parse_brand(val):
    """解析品牌字段 → (是否有自主品牌, 品牌名称)"""
    if not val:
        return '', ''
    val = str(val).strip()
    if val in ('否', '无', 'N', 'NO', '/'):
        return '否', ''
    # 有品牌名称
    return '是', val


# ==================== 读取源数据 ====================

def read_source_data():
    """读取新表 Sheet 1.1 的所有企业数据"""
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb[SOURCE_SHEET]

    enterprises = []
    for row_idx in range(3, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=2).value
        if not name or not str(name).strip():
            continue

        e = {}
        e['name'] = clean_name(name)
        e['district'] = safe_str(ws.cell(row=row_idx, column=3).value)
        e['address'] = safe_str(ws.cell(row=row_idx, column=4).value)
        e['enterprise_type'] = safe_str(ws.cell(row=row_idx, column=6).value)
        e['industry'] = safe_str(ws.cell(row=row_idx, column=7).value)
        e['staff_size'] = safe_str(ws.cell(row=row_idx, column=8).value)
        e['revenue'] = safe_str(ws.cell(row=row_idx, column=9).value)
        e['contact_name'] = safe_str(ws.cell(row=row_idx, column=10).value)
        e['contact_phone'] = safe_str(ws.cell(row=row_idx, column=11).value)
        e['website'] = safe_str(ws.cell(row=row_idx, column=12).value)
        e['products'] = safe_str(ws.cell(row=row_idx, column=13).value)
        e['application'] = safe_str(ws.cell(row=row_idx, column=14).value)
        e['certification'] = safe_str(ws.cell(row=row_idx, column=15).value)
        e['brand'] = safe_str(ws.cell(row=row_idx, column=16).value)
        e['patent'] = safe_str(ws.cell(row=row_idx, column=17).value)
        e['sales_market'] = safe_str(ws.cell(row=row_idx, column=18).value)
        e['local_procurement'] = safe_str(ws.cell(row=row_idx, column=20).value)
        e['automation'] = safe_str(ws.cell(row=row_idx, column=21).value)
        e['capacity'] = safe_str(ws.cell(row=row_idx, column=22).value)
        e['logistics'] = safe_str(ws.cell(row=row_idx, column=24).value)
        e['trade_mode'] = safe_str(ws.cell(row=row_idx, column=25).value)
        e['domestic_ecommerce'] = safe_str(ws.cell(row=row_idx, column=26).value)
        e['import_export_license'] = safe_str(ws.cell(row=row_idx, column=27).value)
        e['customs_mode'] = safe_str(ws.cell(row=row_idx, column=28).value)
        e['trade_team_mode'] = safe_str(ws.cell(row=row_idx, column=29).value)
        e['trade_data'] = safe_str(ws.cell(row=row_idx, column=30).value)
        e['has_cross_border'] = safe_str(ws.cell(row=row_idx, column=31).value)
        e['cross_border_platform'] = safe_str(ws.cell(row=row_idx, column=32).value)
        e['cross_border_ratio'] = safe_str(ws.cell(row=row_idx, column=33).value)
        e['cross_border_logistics'] = safe_str(ws.cell(row=row_idx, column=34).value)
        e['payment_settlement'] = safe_str(ws.cell(row=row_idx, column=35).value)
        e['target_markets'] = safe_str(ws.cell(row=row_idx, column=36).value)
        e['cross_border_team_size'] = safe_str(ws.cell(row=row_idx, column=37).value)
        e['pain_points'] = safe_str(ws.cell(row=row_idx, column=38).value)
        e['has_policy_support'] = safe_str(ws.cell(row=row_idx, column=39).value)
        # 列 40–41 曾为 desired_support / cooperation_demands，已下线，不再写入导出 JSON
        e['competition_position'] = safe_str(ws.cell(row=row_idx, column=42).value)
        e['competitor'] = safe_str(ws.cell(row=row_idx, column=43).value)
        e['risks'] = safe_str(ws.cell(row=row_idx, column=44).value)
        e['transformation_willingness'] = safe_str(ws.cell(row=row_idx, column=45).value)
        # Col 46-56: 跨境需求细分（暂不映射）
        e['service_rating'] = safe_str(ws.cell(row=row_idx, column=57).value)
        e['investment_rating'] = safe_str(ws.cell(row=row_idx, column=58).value)
        e['incubation_rating'] = safe_str(ws.cell(row=row_idx, column=59).value)
        e['brand_rating'] = safe_str(ws.cell(row=row_idx, column=60).value)
        e['training_rating'] = safe_str(ws.cell(row=row_idx, column=61).value)
        e['overall_rating'] = safe_str(ws.cell(row=row_idx, column=62).value)
        e['additional_notes'] = safe_str(ws.cell(row=row_idx, column=63).value)
        e['survey_date'] = safe_str(ws.cell(row=row_idx, column=64).value)
        e['survey_staff'] = safe_str(ws.cell(row=row_idx, column=65).value)
        e['transformation_willingness2'] = safe_str(ws.cell(row=row_idx, column=66).value)
        e['investment_willingness'] = safe_str(ws.cell(row=row_idx, column=67).value)
        e['benchmark_possibility'] = safe_str(ws.cell(row=row_idx, column=68).value)
        e['additional_notes2'] = safe_str(ws.cell(row=row_idx, column=69).value)

        enterprises.append(e)

    wb.close()
    return enterprises


# ==================== 样式 ====================

HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
HINT_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
HINT_FONT = Font(name='微软雅黑', size=10, color='666666', italic=True)
DATA_FONT = Font(name='微软雅黑', size=10)
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)


def style_header(ws, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def style_hint(ws, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=2, column=col)
        cell.fill = HINT_FILL
        cell.font = HINT_FONT
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def style_data(ws, row_count, col_count):
    for row in range(3, row_count + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = THIN_BORDER


# ==================== 生成各Sheet ====================

def build_basic_info_sheet(wb, enterprises):
    """Sheet1: 企业基本信息"""
    ws = wb.create_sheet('企业基本信息')
    headers = ['企业ID', '企业名称', '统一社会信用代码', '所属区域', '详细地址',
               '所属行业', '企业类型', '人员规模', '官网', '国内营收',
               '跨境营收', '企业来源', '是否有自主品牌', '品牌名称', '漏斗阶段']
    widths = [10, 35, 25, 12, 40, 15, 12, 15, 25, 15, 15, 10, 15, 20, 12]

    # 表头
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = widths[i-1]

    # 提示行
    hints = ['', '【请勿修改企业ID】', '18位', '如：武进区', '自由填写',
             '按系统行业分类填写', '生产型/贸易型/工贸一体', '如：50-200人', '自由填写',
             '如：1000-5000万', '', '调研', '是/否', '自由填写', '']
    for i, h in enumerate(hints, 1):
        ws.cell(row=2, column=i, value=h)

    # 数据行
    for idx, e in enumerate(enterprises):
        row = idx + 3
        has_brand, brand_name = parse_brand(e['brand'])
        district = DISTRICT_MAP.get(e['district'], e['district'])
        etype = TYPE_MAP.get(e['enterprise_type'], e['enterprise_type'])
        revenue = parse_revenue_range(e['revenue'])

        ws.cell(row=row, column=1, value=None)  # 企业ID留空
        ws.cell(row=row, column=2, value=e['name'])
        ws.cell(row=row, column=3, value='')  # 信用代码（新表无）
        ws.cell(row=row, column=4, value=district)
        ws.cell(row=row, column=5, value=e['address'])
        ws.cell(row=row, column=6, value=normalize_industry(e['industry']))
        ws.cell(row=row, column=7, value=etype)
        ws.cell(row=row, column=8, value=e['staff_size'])
        ws.cell(row=row, column=9, value=e['website'])
        ws.cell(row=row, column=10, value=revenue)
        ws.cell(row=row, column=11, value='')  # 跨境营收（新表无单独字段）
        ws.cell(row=row, column=12, value='调研')  # 企业来源
        ws.cell(row=row, column=13, value=has_brand)
        ws.cell(row=row, column=14, value=brand_name)
        ws.cell(row=row, column=15, value='')  # 漏斗阶段

    style_header(ws, len(headers))
    style_hint(ws, len(headers))
    style_data(ws, len(enterprises) + 2, len(headers))
    # 隐藏企业ID列
    ws.column_dimensions['A'].hidden = True


def build_contact_sheet(wb, enterprises):
    """Sheet2: 联系人信息"""
    ws = wb.create_sheet('联系人信息')
    headers = ['企业ID', '企业名称', '联系人姓名', '联系电话', '职位',
               '邮箱', '微信', '是否主要联系人', '备注']
    widths = [10, 35, 15, 18, 15, 25, 18, 15, 30]

    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = widths[i-1]

    hints = ['', '【请勿修改企业ID】', '必填', '手机号', '自由填写',
             '自由填写', '自由填写', '是/否', '自由填写']
    for i, h in enumerate(hints, 1):
        ws.cell(row=2, column=i, value=h)

    row = 3
    for e in enterprises:
        if not e['contact_name']:
            continue
        ws.cell(row=row, column=1, value=None)
        ws.cell(row=row, column=2, value=e['name'])
        ws.cell(row=row, column=3, value=e['contact_name'])
        ws.cell(row=row, column=4, value=e['contact_phone'])
        ws.cell(row=row, column=5, value='')
        ws.cell(row=row, column=6, value='')
        ws.cell(row=row, column=7, value='')
        ws.cell(row=row, column=8, value='是')
        ws.cell(row=row, column=9, value='')
        row += 1

    style_header(ws, len(headers))
    style_hint(ws, len(headers))
    style_data(ws, row - 1, len(headers))
    ws.column_dimensions['A'].hidden = True


def build_product_sheet(wb, enterprises):
    """Sheet3: 产品信息"""
    ws = wb.create_sheet('产品信息')
    headers = ['企业ID', '企业名称', '产品名称', '产品品类', '认证资质',
               '主要销售区域', '主要销售国家', '年销售额', '原材料本地采购比例',
               '设备自动化程度', '年产能', '物流合作方']
    widths = [10, 35, 20, 15, 40, 35, 35, 15, 18, 20, 15, 30]

    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = widths[i-1]

    hints = ['', '【请勿修改企业ID】', '必填', '按系统品类填写', '多选用顿号分隔',
             '多选用顿号分隔', '多选用顿号分隔', '自由填写', '如：80%',
             '低/中/高/很高', '自由填写', '多选用顿号分隔']
    for i, h in enumerate(hints, 1):
        ws.cell(row=2, column=i, value=h)

    row = 3
    for e in enterprises:
        if not e['products']:
            continue
        # 拆分多个产品（逗号、顿号、分号分隔）
        product_names = re.split(r'[,，、;；]', e['products'])
        for pname in product_names:
            pname = pname.strip()
            if not pname:
                continue
            ws.cell(row=row, column=1, value=None)
            ws.cell(row=row, column=2, value=e['name'])
            ws.cell(row=row, column=3, value=pname)
            ws.cell(row=row, column=4, value='')  # 品类需人工匹配
            ws.cell(row=row, column=5, value=e['certification'])
            ws.cell(row=row, column=6, value=e['sales_market'])
            ws.cell(row=row, column=7, value='')
            ws.cell(row=row, column=8, value='')
            ws.cell(row=row, column=9, value=e['local_procurement'])
            ws.cell(row=row, column=10, value=e['automation'])
            ws.cell(row=row, column=11, value=e['capacity'])
            ws.cell(row=row, column=12, value=e['logistics'])
            row += 1

    style_header(ws, len(headers))
    style_hint(ws, len(headers))
    style_data(ws, row - 1, len(headers))
    ws.column_dimensions['A'].hidden = True


def build_trade_sheet(wb, enterprises):
    """Sheet4: 外贸信息"""
    ws = wb.create_sheet('外贸信息')
    headers = ['企业ID', '企业名称', '主要销售区域', '主要销售国家', '外贸模式',
               '是否有进出口资质', '报关申报主体模式', '外贸团队模式', '外贸团队人数',
               '是否有国内电商经验', '上年外贸营业额(万元)', '上上年外贸营业额(万元)',
               '增长市场', '下降市场', '增长模式', '下降模式',
               '增长品类', '下降品类', '增长原因', '下降原因']
    widths = [10, 35, 35, 35, 25, 18, 18, 25, 12, 18, 18, 18,
              30, 30, 30, 30, 30, 30, 35, 35]

    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = widths[i-1]

    hints = ['', '【请勿修改企业ID】', '多选用顿号分隔', '多选用顿号分隔',
             '如：0110/1039/9610等', '是/否', '自由填写', '自营团队/外包团队/混合模式/无专职团队',
             '填数字', '是/否', '填数字(万元)', '填数字(万元)',
             '格式：市场名 +变化率', '格式：市场名 -变化率',
             '格式：模式名 +变化率', '格式：模式名 -变化率',
             '格式：品类名 +变化率', '格式：品类名 -变化率',
             '多个用顿号分隔', '多个用顿号分隔']
    for i, h in enumerate(hints, 1):
        ws.cell(row=2, column=i, value=h)

    for idx, e in enumerate(enterprises):
        row = idx + 3
        ws.cell(row=row, column=1, value=None)
        ws.cell(row=row, column=2, value=e['name'])
        ws.cell(row=row, column=3, value=e['sales_market'])
        ws.cell(row=row, column=4, value='')
        ws.cell(row=row, column=5, value=e['trade_mode'])
        ws.cell(row=row, column=6, value=normalize_yes_no(e['import_export_license']))
        ws.cell(row=row, column=7, value=e['customs_mode'])
        ws.cell(row=row, column=8, value=e['trade_team_mode'])
        ws.cell(row=row, column=9, value='')  # 团队人数（新表无）
        ws.cell(row=row, column=10, value=normalize_yes_no(e['domestic_ecommerce']))
        ws.cell(row=row, column=11, value='')  # 上年营业额（新表只有笼统数据）
        ws.cell(row=row, column=12, value='')
        # 增长/下降字段（新表无）
        for col in range(13, 21):
            ws.cell(row=row, column=col, value='')

    style_header(ws, len(headers))
    style_hint(ws, len(headers))
    style_data(ws, len(enterprises) + 2, len(headers))
    ws.column_dimensions['A'].hidden = True


def build_cross_border_sheet(wb, enterprises):
    """Sheet5: 跨境电商信息"""
    ws = wb.create_sheet('跨境电商信息')
    headers = ['企业ID', '企业名称', '是否开展跨境电商', '跨境平台', '跨境业务占比',
               '跨境物流模式', '支付结算方式', '跨境电商团队规模', '是否在用ERP',
               '跨境转型意愿', '愿意投入转型程度', '目标市场及占比']
    widths = [10, 35, 18, 45, 15, 35, 35, 18, 15, 15, 18, 40]

    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = widths[i-1]

    hints = ['', '【请勿修改企业ID】', '是/否', '多选用顿号分隔', '如：25%',
             '多选用顿号分隔', '多选用顿号分隔', '填数字', '是/否',
             '高/中/低', '高/中/低', '格式：市场名 占比%']
    for i, h in enumerate(hints, 1):
        ws.cell(row=2, column=i, value=h)

    for idx, e in enumerate(enterprises):
        row = idx + 3
        # 转型意愿：优先用 Col66，否则用 Col45
        willingness = e['transformation_willingness2'] or e['transformation_willingness']
        # 标准化转型意愿为 高/中/低
        will_val = ''
        if willingness:
            w = willingness.strip()
            if w in ('高', '中', '低'):
                will_val = w
            elif '有' in w or '强' in w or '高' in w:
                will_val = '高'
            elif '一般' in w or '中' in w:
                will_val = '中'
            elif '无' in w or '低' in w or '否' in w:
                will_val = '低'

        invest_val = ''
        if e['investment_willingness']:
            iv = e['investment_willingness'].strip()
            if iv in ('高', '中', '低'):
                invest_val = iv

        ws.cell(row=row, column=1, value=None)
        ws.cell(row=row, column=2, value=e['name'])
        ws.cell(row=row, column=3, value=normalize_yes_no(e['has_cross_border']))
        ws.cell(row=row, column=4, value=e['cross_border_platform'])
        ws.cell(row=row, column=5, value=e['cross_border_ratio'])
        ws.cell(row=row, column=6, value=e['cross_border_logistics'])
        ws.cell(row=row, column=7, value=e['payment_settlement'])
        ws.cell(row=row, column=8, value=e['cross_border_team_size'])
        ws.cell(row=row, column=9, value='')  # ERP（新表无）
        ws.cell(row=row, column=10, value=will_val)
        ws.cell(row=row, column=11, value=invest_val)
        ws.cell(row=row, column=12, value=e['target_markets'])

    style_header(ws, len(headers))
    style_hint(ws, len(headers))
    style_data(ws, len(enterprises) + 2, len(headers))
    ws.column_dimensions['A'].hidden = True


def build_cooperation_sheet(wb, enterprises):
    """Sheet6: 合作与政策信息"""
    ws = wb.create_sheet('合作与政策信息')
    headers = ['企业ID', '企业名称', '企业服务合作(1-5星)', '招商入驻合作(1-5星)',
               '孵化转型合作(1-5星)', '品牌营销合作(1-5星)', '人才培训合作(1-5星)',
               '跨境整体方案(1-5星)', '标杆企业可能性(%)', '是否享受过政策支持',
               '已享受政策', '调研日期', '调研人员', '行业竞争地位',
               '竞争地位描述', '当前面临风险', '风险详细描述', '补充说明', '建议事项']
    widths = [10, 35, 18, 18, 18, 18, 18, 18, 18, 18,
              40, 15, 15, 15, 30, 40, 35, 35, 35]

    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = widths[i-1]

    hints = ['', '【请勿修改企业ID】', '1-5', '1-5', '1-5', '1-5', '1-5', '1-5',
             '0-100', '是/否', '多选用顿号分隔', '如：2024-01-15', '自由填写',
             '头部企业/中型企业/初创企业', '自由填写', '多选用顿号分隔',
             '自由填写', '自由填写', '自由填写']
    for i, h in enumerate(hints, 1):
        ws.cell(row=2, column=i, value=h)

    for idx, e in enumerate(enterprises):
        row = idx + 3

        def safe_rating(val):
            if not val:
                return ''
            try:
                v = int(float(str(val).strip()))
                return str(v) if 1 <= v <= 5 else ''
            except (ValueError, TypeError):
                return ''

        def safe_benchmark(val):
            if not val:
                return ''
            try:
                v = int(float(str(val).strip().replace('%', '')))
                return str(v) if 0 <= v <= 100 else ''
            except (ValueError, TypeError):
                return ''

        # 合并补充说明
        notes_parts = [e['additional_notes'], e['additional_notes2'], e['pain_points']]
        notes = '、'.join([n for n in notes_parts if n])

        ws.cell(row=row, column=1, value=None)
        ws.cell(row=row, column=2, value=e['name'])
        ws.cell(row=row, column=3, value=safe_rating(e['service_rating']))
        ws.cell(row=row, column=4, value=safe_rating(e['investment_rating']))
        ws.cell(row=row, column=5, value=safe_rating(e['incubation_rating']))
        ws.cell(row=row, column=6, value=safe_rating(e['brand_rating']))
        ws.cell(row=row, column=7, value=safe_rating(e['training_rating']))
        ws.cell(row=row, column=8, value=safe_rating(e['overall_rating']))
        ws.cell(row=row, column=9, value=safe_benchmark(e['benchmark_possibility']))
        ws.cell(row=row, column=10, value=normalize_yes_no(e['has_policy_support']))
        ws.cell(row=row, column=11, value='')  # 已享受政策（新表无具体列表）
        ws.cell(row=row, column=12, value=e['survey_date'])
        ws.cell(row=row, column=13, value=e['survey_staff'])
        ws.cell(row=row, column=14, value=e['competition_position'])
        ws.cell(row=row, column=15, value='')  # 竞争地位描述
        ws.cell(row=row, column=16, value=e['risks'])
        ws.cell(row=row, column=17, value='')
        ws.cell(row=row, column=18, value=notes)
        ws.cell(row=row, column=19, value='')

    style_header(ws, len(headers))
    style_hint(ws, len(headers))
    style_data(ws, len(enterprises) + 2, len(headers))
    ws.column_dimensions['A'].hidden = True


# ==================== 主函数 ====================

def main():
    print(f'读取源文件: {INPUT_FILE}')
    enterprises = read_source_data()
    print(f'读取到 {len(enterprises)} 条企业数据')

    # 去重（按企业名称）
    seen = set()
    unique = []
    dup_count = 0
    for e in enterprises:
        if e['name'] not in seen:
            seen.add(e['name'])
            unique.append(e)
        else:
            dup_count += 1
    if dup_count > 0:
        print(f'去重: 移除 {dup_count} 条重复企业，剩余 {len(unique)} 条')
    enterprises = unique

    # 统计有效数据
    has_contact = sum(1 for e in enterprises if e['contact_name'])
    has_product = sum(1 for e in enterprises if e['products'])
    has_trade = sum(1 for e in enterprises if e['trade_mode'])
    has_cb = sum(1 for e in enterprises if e['has_cross_border'])
    print(f'有联系人: {has_contact}, 有产品: {has_product}, 有外贸: {has_trade}, 有跨境: {has_cb}')

    # 生成输出文件
    wb = openpyxl.Workbook()
    # 删除默认Sheet
    wb.remove(wb.active)

    build_basic_info_sheet(wb, enterprises)
    build_contact_sheet(wb, enterprises)
    build_product_sheet(wb, enterprises)
    build_trade_sheet(wb, enterprises)
    build_cross_border_sheet(wb, enterprises)
    build_cooperation_sheet(wb, enterprises)

    wb.save(OUTPUT_FILE)
    print(f'\n转换完成! 输出文件: {OUTPUT_FILE}')
    print(f'共 {len(enterprises)} 条企业数据，6个Sheet')
    print('\n注意事项:')
    print('1. 企业ID列为空，导入时系统会自动创建新企业')
    print('2. 产品品类列为空，需人工匹配系统品类或导入后在系统中修改')
    print('3. 部分选项字段（如人员规模、营收区间）可能需要与系统选项对齐')
    print('4. 建议先导入少量数据测试，确认无误后再全量导入')


if __name__ == '__main__':
    main()
