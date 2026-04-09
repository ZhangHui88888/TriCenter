# -*- coding: utf-8 -*-
"""
清理"企业服务数据新表.xlsm" Sheet 1.1 中公司地址的市/区前缀

由于所有公司默认是常州市，且区信息已在前一列，
本脚本去除地址中的"常州市"、"XX区"等冗余前缀。

用法: python scripts/clean_address.py
输入: docs/data/企业服务数据新表.xlsm (Sheet 1.1)
输出: docs/data/企业服务数据新表_地址已清理.xlsx
"""

import openpyxl
import re
import os

INPUT_FILE = os.path.join('docs', 'data', '企业服务数据新表.xlsm')
OUTPUT_FILE = os.path.join('docs', 'data', '企业服务数据新表_地址已清理.xlsx')
SOURCE_SHEET = '1.1 常州市产业带调研'
ADDRESS_COLUMN = 4  # 第四列是地址（第三列是区域）

# 需要去除的市/区前缀模式
CITY_DISTRICT_PATTERNS = [
    r'^常州市',
    r'^江苏省常州市',
    r'^江苏常州市',
    r'^江苏省',
    # 常州各区
    r'^新北区',
    r'^武进区',
    r'^天宁区',
    r'^钟楼区',
    r'^经开区',
    r'^经济开发区',
    r'^金坛区',
    r'^溧阳市',
    # 简称
    r'^新北',
    r'^武进',
    r'^天宁',
    r'^钟楼',
    r'^经开',
    r'^金坛',
    r'^溧阳',
]


def clean_address(address):
    """去除地址中的市/区前缀"""
    if not address:
        return address
    
    original = str(address).strip()
    cleaned = original
    
    # 循环去除前缀（可能有多个，如"常州市新北区"）
    changed = True
    while changed:
        changed = False
        for pattern in CITY_DISTRICT_PATTERNS:
            new_val = re.sub(pattern, '', cleaned, count=1)
            if new_val != cleaned:
                cleaned = new_val.strip()
                changed = True
                break
    
    return cleaned


def main():
    print(f'读取源文件: {INPUT_FILE}')
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb[SOURCE_SHEET]
    
    print(f'工作表: {SOURCE_SHEET}')
    print(f'总行数: {ws.max_row}')
    
    # 统计
    total = 0
    modified = 0
    
    # 从第3行开始（跳过表头和提示行）
    for row_idx in range(3, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=ADDRESS_COLUMN)
        original = cell.value
        
        if not original:
            continue
        
        total += 1
        cleaned = clean_address(original)
        
        if cleaned != str(original).strip():
            print(f'行 {row_idx}: "{original}" -> "{cleaned}"')
            cell.value = cleaned
            modified += 1
    
    print(f'\n处理完成:')
    print(f'  总地址数: {total}')
    print(f'  已修改: {modified}')
    print(f'  未修改: {total - modified}')
    
    # 保存为新文件（避免覆盖原文件）
    wb.save(OUTPUT_FILE)
    print(f'\n输出文件: {OUTPUT_FILE}')
    wb.close()


if __name__ == '__main__':
    main()
