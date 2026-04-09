"""
精确匹配 Excel 产品描述到系统品类名称，并写回 Excel。
匹配策略：
  1. 拆词后先精确匹配系统品类名
  2. 再查 MANUAL_MAP（别名/长描述 -> 品类名）
  3. 匹配到的品类名写入 Excel 的"匹配品类"列
"""
import re
import openpyxl
from copy import copy

# ============================================================
# 1. 从 SQL 提取所有品类 (id, name, level, path, parent_id)
# ============================================================
categories = {}  # id -> {name, level, path, parent_id}

in_product_categories = False
with open(r"D:\github\funNovels\TriCenter\docs\sql\tricenter_init.sql", encoding="utf-8") as f:
    for line in f:
        # 检测 INSERT INTO 语句切换表名上下文
        if "INSERT INTO product_categories" in line:
            in_product_categories = True
        elif "INSERT INTO " in line and "product_categories" not in line:
            in_product_categories = False

        if not in_product_categories:
            continue

        m = re.match(
            r"\s*\((\d+),\s*(\d+),\s*'([^']+)',\s*(\d+),\s*'([^']+)',\s*(\d+)\)",
            line,
        )
        if m:
            cid = int(m.group(1))
            pid = int(m.group(2))
            name = m.group(3)
            level = int(m.group(4))
            path = m.group(5)
            categories[cid] = {
                "name": name,
                "level": level,
                "path": path,
                "parent_id": pid,
            }

def build_full_path(cid):
    """根据 path 字段构建中文全路径"""
    info = categories.get(cid)
    if not info:
        return ""
    path_ids = info["path"].split("/")
    path_names = []
    for pid_str in path_ids:
        pid = int(pid_str)
        if pid in categories:
            path_names.append(categories[pid]["name"])
    return " > ".join(path_names)

# 构建 name -> [{id, level, full_path}]
name_map = {}
for cid, info in categories.items():
    fp = build_full_path(cid)
    if info["name"] not in name_map:
        name_map[info["name"]] = []
    name_map[info["name"]].append(
        {"id": cid, "level": info["level"], "full_path": fp}
    )

# 统计
l1 = sum(1 for n, v in name_map.items() if any(x["level"] == 1 for x in v))
l2 = sum(1 for n, v in name_map.items() if any(x["level"] == 2 for x in v))
l3 = sum(1 for n, v in name_map.items() if any(x["level"] == 3 for x in v))
print(f"系统品类总数: {len(categories)}")
print(f"  一级: {l1}, 二级: {l2}, 三级: {l3}")
print(f"  品类名称(去重): {len(name_map)}")
print()

# ============================================================
# 2. MANUAL_MAP: 别名/长描述 -> 品类名（必须已存在于系统品类中）
#    None 表示噪音/非产品词，跳过
# ============================================================
MANUAL_MAP = {
    "整体卫浴房（玻璃钢材质）": "整体卫浴",
    "各类T恤": "T恤",
    "外套和裤子": "外套",
    "牛仔类休闲服装": "牛仔休闲服装",
    "一次性使用注射器": "一次性注射器",
    "一次性使用无菌注射器 带针": "一次性注射器",
    "一次性使用注射器一次性使用活体取样钳": "一次性注射器",
    "扩阴器": "扩张器",
    "导管类": "导管",
    "尿袋喂食袋": "尿袋",
    "各类医疗器械的生产": "医疗设备",
    "制药化工干燥设备": "制药干燥设备",
    "连接器等": "连接器",
    "电子标签等": "电子标签",
    "环网柜核心部件": "环网柜部件",
    "液压管件出口": "液压管件",
    "齿轮以及传动部分": "齿轮",
    "微特电机及配件": "微特电机",
    "伺服装置及微电机": "伺服装置",
    "农业机械研发": "农业机械",
    "泵及真空设备销售": "真空设备",
    "高压无刷电机及变频控制器的研发与制造": "无刷电机",
    "动力系统成套解决方案（无刷角磨机": "无刷角磨机",
    "永磁高速深井水泵等）": "深井水泵",
    "油泵及水泵（覆盖1kW-500kW全系列）": "油泵及水泵",
    "汽车衡 小地磅": "汽车衡",
    "抗震领域新材料": "抗震新材料",
    "地膜及塑料制品": "地膜",
    "摩托车头盔及骑行防护用品": "摩托车头盔",
    "刷子生产": "刷子",
    "飞机用坐具的零件": "飞机坐具零件",
    "精密轴承的制造与销售": "精密轴承",
    "高品质精密金属切削刀具": "精密切削刀具",
    "超高精密微纳3D打印系统": "微纳3D打印系统",
    "HRM EAM": "HRM",
    "针对企业人力资源管理咨询": "人力资源管理咨询",
    "针对企业数智化产品开发": "数智化产品开发",
    "针对出海企业服务产品": "出海企业服务",
    "医院运营和质控平台": "医院运营平台",
    "计算机系统集成": "系统集成",
    "计算机软硬件开发": "软硬件开发",
    "深度电商供应链管理": "电商供应链管理",
    "一站式电商服务平台（代运营": "电商代运营",
    "网页和网站设计": "网站设计",
    "常州本地组装。盐城工厂做光伏产品加工": "光伏设备",
    "汽车零部件": "汽车配件",
    "汽配": "汽车配件",
    "汽配产品": "汽车配件",
    # 噪音词
    "销售": None, "研发": None, "生产": None, "等": None,
    "洛杉矶海外仓7千平方": None, "线下实体店100平方": None,
    "3个": None, "核心部件外部采购": None, "设计": None,
    "客服": None, "推广等）": None,
}

# ============================================================
# 3. 读取 Excel，拆词精确匹配 + MANUAL_MAP
# ============================================================
excel_path = r"D:\github\funNovels\TriCenter\docs\data\企业服务数据新表_行业已匹配.xlsx"
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

SPLIT_RE = re.compile(r"[,，、；;/]")

# 在第14列写入表头"匹配品类"
MATCH_COL = 14
ws.cell(2, MATCH_COL, "匹配品类")

matched_count = 0
unmatched_enterprises = []
total = 0

for r in range(3, ws.max_row + 1):
    name = ws.cell(r, 2).value
    if not name or str(name).strip() == "":
        continue
    name = str(name).strip()

    product = ws.cell(r, 13).value
    product_str = str(product).strip() if product else ""
    if not product_str or product_str == "None":
        continue

    total += 1
    tokens = [t.strip() for t in SPLIT_RE.split(product_str) if t.strip()]

    row_matched_names = []  # 匹配到的品类名（去重、保序）
    row_unmatched = []

    for token in tokens:
        # 策略1: 精确匹配品类名
        if token in name_map:
            if token not in row_matched_names:
                row_matched_names.append(token)
            continue
        # 策略2: MANUAL_MAP 别名映射
        if token in MANUAL_MAP:
            mapped = MANUAL_MAP[token]
            if mapped is not None and mapped in name_map:
                if mapped not in row_matched_names:
                    row_matched_names.append(mapped)
            continue
        row_unmatched.append(token)

    if row_matched_names:
        matched_count += 1
        # 构建输出: 品类名1, 品类名2
        ws.cell(r, MATCH_COL, ", ".join(row_matched_names))
    else:
        unmatched_enterprises.append((name, product_str, row_unmatched))

# 保存
output_path = r"D:\github\funNovels\TriCenter\docs\data\企业服务数据新表_品类已匹配.xlsx"
wb.save(output_path)
wb.close()

# ============================================================
# 4. 输出统计
# ============================================================
print("=" * 70)
print(f"企业总数（有产品描述）: {total}")
print(f"匹配成功: {matched_count} ({matched_count*100//total}%)")
print(f"未匹配: {len(unmatched_enterprises)}")
print("=" * 70)

if unmatched_enterprises:
    print(f"\n【仍未匹配的企业（{len(unmatched_enterprises)} 家）】")
    for i, (name, product, tokens) in enumerate(unmatched_enterprises, 1):
        print(f"  {i}. {name}: {product}")
        print(f"     未匹配拆词: {tokens}")

print(f"\n结果已写入: {output_path}")
