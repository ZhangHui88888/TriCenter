# -*- coding: utf-8 -*-
"""
匹配"企业服务数据新表_地址已清理.xlsx" Sheet 1.1 中的所属行业
转换为系统可选的标准行业分类

系统标准行业：园艺制品、电动工具、汽车零部件、家居建材、机械设备、纺织服装、电子产品、其他

用法: python scripts/match_industry.py
"""

import openpyxl
import os
from collections import Counter

INPUT_FILE = os.path.join('docs', 'data', '企业服务数据新表_地址已清理.xlsx')
OUTPUT_FILE = os.path.join('docs', 'data', '企业服务数据新表_行业已匹配.xlsx')
SOURCE_SHEET = '1.1 常州市产业带调研'
INDUSTRY_COLUMN = 7  # 第7列是所属行业

# 系统标准行业
STANDARD_INDUSTRIES = [
    '园艺制品',
    '电动工具',
    '汽车零部件',
    '家居建材',
    '机械设备',
    '纺织服装',
    '电子产品',
    '其他',
]

# 行业映射规则：原始值 → 标准行业
INDUSTRY_MAP = {
    # 汽车零部件
    '汽配': '汽车零部件',
    '汽车配件': '汽车零部件',
    '汽车部件': '汽车零部件',
    '汽车配套工具': '汽车零部件',
    '汽车制造': '汽车零部件',
    '汽车零部件': '汽车零部件',
    '汽车零部件，机械零部件，轴承': '汽车零部件',
    
    # 纺织服装
    '纺织': '纺织服装',
    '纺织品': '纺织服装',
    '纺织业': '纺织服装',
    '纺织服装类': '纺织服装',
    '纺织科技': '纺织服装',
    '纺机': '纺织服装',
    '针纺织品': '纺织服装',
    '服装': '纺织服装',
    '服装制造': '纺织服装',
    '服装贸易进出口': '纺织服装',
    '服装贸易': '纺织服装',
    '纺织服装': '纺织服装',
    
    # 机械设备
    '机械': '机械设备',
    '机械设备': '机械设备',
    '机械设备制造': '机械设备',
    '干燥设备': '机械设备',
    '干燥业': '机械设备',
    '通用设备': '机械设备',
    '35专用设备制造业': '机械设备',
    '专用设备制造业': '机械设备',
    '环保设备 / 净化科技': '机械设备',
    '先进制造 / 电机及控制器制造': '机械设备',
    '机电行业': '机械设备',
    '医疗器械': '机械设备',
    '医疗': '机械设备',
    '医疗行业': '机械设备',
    '制药装备': '机械设备',
    '通信设备': '机械设备',
    
    # 电子产品
    '电子': '电子产品',
    '电子电工': '电子产品',
    '电子机械': '电子产品',
    '电机行业': '电子产品',
    '3c': '电子产品',
    '3c电子': '电子产品',
    '电子产品': '电子产品',
    
    # 家居建材（含五金）
    '五金': '家居建材',
    '五金工具': '家居建材',
    '五金工具制造': '家居建材',
    '建材': '家居建材',
    '建筑': '家居建材',
    '建筑工程': '家居建材',
    '装饰材料': '家居建材',
    '钣金制造': '家居建材',
    '高压配电': '家居建材',
    '齿轮': '家居建材',
    '家居': '家居建材',
    '家具': '家居建材',
    '家居园艺': '家居建材',
    '智能家居': '家居建材',
    '厨具': '家居建材',
    '日用百货': '家居建材',
    '办公用品': '家居建材',
    '刀具': '家居建材',
    '建筑材料 / 化工产品 / 合成材料': '家居建材',
    '建筑材料': '家居建材',
    '家居建材': '家居建材',
    '五金建材': '家居建材',
    
    # 园艺制品
    '园艺': '园艺制品',
    '园艺制品': '园艺制品',
    
    # 电动工具
    '电动工具': '电动工具',
    '扳机': '电动工具',
    
    # 照明灯具 → 家居建材
    '灯具': '家居建材',
    '灯具照明': '家居建材',
    '照明': '家居建材',
    '照明灯具': '家居建材',
    
    # 化工/新能源/新材料 → 其他
    '化工': '其他',
    '塑料包装': '其他',
    '新材料': '其他',
    '光伏': '其他',
    '光伏行业': '其他',
    '新能源': '其他',
    '化工材料': '其他',
    '新能源航空动力系统 / 电机制造': '其他',
    
    # 贸易/电商/服务类 → 其他
    '贸易': '其他',
    '工贸一体': '其他',
    '卖家': '其他',
    '批发外贸': '其他',
    '批发商': '其他',
    '批发': '其他',
    '批发业': '其他',
    '批发和零售业': '其他',
    '批发零售业': '其他',
    '多品类': '其他',
    '电商': '其他',
    '跨境卖家': '其他',
    '跨境电商': '其他',
    '互联网销售': '其他',
    '进出口': '其他',
    '艺术品进出口': '其他',
    '技术进出口': '其他',
    '综合贸易': '其他',
    '批发零售': '其他',
    '电商运营': '其他',
    '进出口代理': '其他',
    '互联网 / 电子商务': '其他',
    
    # 信息技术/软件 → 其他
    '软件': '其他',
    '信息工程': '其他',
    '信息传输软件业': '其他',
    '技术服务': '其他',
    '人工智能': '其他',
    '智能机器人': '其他',
    '高新技术业': '其他',
    '互联网信息': '其他',
    '信息技术': '其他',
    '计算机信息技术服务 / 系统集成': '其他',
    
    # 物流/金融/营销 → 其他
    '海路运输': '其他',
    '金融': '其他',
    '广告': '其他',
    '展示器材': '其他',
    '物流仓储': '其他',
    '金融服务': '其他',
    '营销推广': '其他',
    
    # 其他制造/服务 → 其他
    '制造': '其他',
    '制造业': '其他',
    '工业制造': '其他',
    '工业品': '其他',
    '生产制造': '其他',
    '生产加工': '其他',
    '印刷': '其他',
    '文化': '其他',
    '轻工业': '其他',
    '成人用品': '其他',
    '其他制造': '其他',
    '咨询服务': '其他',
    '专业设计服务': '其他',
    '服务业': '其他',
    '箱包': '其他',
    '箱包皮具': '其他',
    
    # 空值和占位符
    '.': '',
    '': '',
    
    # 补充未匹配项
    '货物进出口': '其他',
    '文教体育用品批发': '其他',
    '计算机信息技术服务 ， 系统集成': '其他',
    '建筑材料 ， 化工产品 ， 合成材料': '家居建材',
    '集成电路设计': '电子产品',
    '互联网零售': '其他',
    '航空服务': '其他',
    '专业技术服务业': '其他',
    'IT技术服务': '其他',
    '物联网': '电子产品',
}


def match_industry(val):
    """尝试匹配行业到标准分类"""
    if not val:
        return '', None
    
    val = str(val).strip()
    if val in ('.', '', '/'):
        return '', None
    
    # 直接匹配
    if val in STANDARD_INDUSTRIES:
        return val, None
    
    # 映射匹配
    if val in INDUSTRY_MAP:
        return INDUSTRY_MAP[val], None
    
    # 尝试部分匹配（处理复合名称）
    val_lower = val.lower()
    
    # 关键词匹配
    if any(kw in val for kw in ['汽车', '汽配']):
        return '汽车零部件', None
    if any(kw in val for kw in ['纺织', '服装', '针织']):
        return '纺织服装', None
    if any(kw in val for kw in ['机械', '设备', '制造']):
        return '机械设备', None
    if any(kw in val for kw in ['电子', '电气', '电机']):
        return '电子产品', None
    if any(kw in val for kw in ['五金', '建材', '家居', '家具', '灯具']):
        return '家居建材', None
    if any(kw in val for kw in ['园艺']):
        return '园艺制品', None
    if any(kw in val for kw in ['电动工具']):
        return '电动工具', None
    
    # 无法匹配
    return None, val


def main():
    print(f'读取文件: {INPUT_FILE}')
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb[SOURCE_SHEET]
    
    print(f'工作表: {SOURCE_SHEET}')
    print(f'总行数: {ws.max_row}')
    
    # 统计
    total = 0
    matched = 0
    unmatched_list = []
    unmatched_counter = Counter()
    
    # 从第3行开始
    for row_idx in range(3, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=INDUSTRY_COLUMN)
        original = cell.value
        
        if not original or str(original).strip() in ('.', '', '/'):
            continue
        
        total += 1
        result, unmatched = match_industry(original)
        
        if unmatched:
            # 无法匹配
            enterprise_name = ws.cell(row=row_idx, column=2).value
            unmatched_list.append((row_idx, enterprise_name, original))
            unmatched_counter[original] += 1
        elif result:
            # 匹配成功，更新单元格
            if result != str(original).strip():
                cell.value = result
            matched += 1
        else:
            # 空值，跳过
            pass
    
    print(f'\n========== 匹配结果 ==========')
    print(f'总行业数: {total}')
    print(f'已匹配: {matched}')
    print(f'未匹配: {len(unmatched_list)}')
    
    if unmatched_counter:
        print(f'\n========== 未匹配的行业（按出现次数排序）==========')
        for industry, count in unmatched_counter.most_common():
            print(f'  [{count}次] {industry}')
        
        print(f'\n========== 未匹配的详细记录 ==========')
        for row_idx, name, industry in unmatched_list:
            print(f'  行{row_idx}: {name} - 行业: {industry}')
    
    # 保存
    wb.save(OUTPUT_FILE)
    print(f'\n输出文件: {OUTPUT_FILE}')
    wb.close()


if __name__ == '__main__':
    main()
